from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl


FILENAME = 'formats.xlsx'

workbook = load_workbook(filename=FILENAME)
worksheet = workbook.get_sheet_by_name('abcd')

# Set column widths
worksheet.column_dimensions['A'].width = 56
worksheet.column_dimensions['B'].width = 13
worksheet.column_dimensions['C'].width = 38
worksheet.column_dimensions['D'].width = 17
worksheet.column_dimensions['E'].width = 27
worksheet.column_dimensions['F'].width = 35
worksheet.column_dimensions['G'].width = 16
worksheet.column_dimensions['H'].width = 16
worksheet.column_dimensions['I'].width = 31
worksheet.column_dimensions['J'].width = 31
worksheet.column_dimensions['K'].width = 31
worksheet.column_dimensions['L'].width = 31


# worksheet['A1'].alignment = Alignment(wrapText=True)

# Set wrap text
for column_letter in ['F', 'I', 'J', 'K', 'L']:
    cells = worksheet[column_letter]
    for cell in cells:
        cell.alignment = Alignment(wrapText=True)


# Set header row bold, shaded
# https://stackoverflow.com/questions/46162147/python-openpyxl-change-font-to-bold?rq=1
#
for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    cell = worksheet['{0}{1}'.format(column_letter, 1)]
    cell.font = Font(bold=True)
    shade = openpyxl.styles.colors.Color(rgb='BFBFBF')
    fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=shade)
    cell.fill = fill

# Border
border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
# A - L = chr 65-77
for column_number in range(1,13):
    for row_number in range(1,77):
        cell = worksheet['{0}{1}'.format(chr(column_number + 64), row_number)]
        cell.border = border




# Save and Close
workbook.save(FILENAME)
workbook.close()

print('DONE')
